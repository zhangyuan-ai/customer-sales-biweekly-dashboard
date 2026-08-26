"""Generate dashboard data from two local customer sales reports.

Usage:
  python scripts/update-dashboard-data.py PREVIOUS.xlsx CURRENT.xlsx
  python scripts/update-dashboard-data.py --mode monthly PREVIOUS.xlsx CURRENT.xlsx

The source workbooks stay local. Only aggregated dashboard JSON is written.
"""

from __future__ import annotations

import json
import re
import sys
from calendar import monthrange
from collections import OrderedDict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


REQUIRED_COLUMNS = ("客户公司", "客户名称", "含税金额", "参考毛利额")
REPO_ROOT = Path(__file__).resolve().parents[1]
WEEKLY_DATA_PATH = REPO_ROOT / "app" / "dashboard-data.json"
WEEKLY_HISTORY_PATH = REPO_ROOT / "app" / "dashboard-history.json"
MONTHLY_DATA_PATH = REPO_ROOT / "app" / "monthly-dashboard-data.json"
MONTHLY_HISTORY_PATH = REPO_ROOT / "app" / "monthly-dashboard-history.json"


def number(value: object) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    return float(str(value).replace(",", "").strip())


def period_from_filename(path: Path, mode: str) -> dict[str, object]:
    dates = re.findall(r"20\d{2}-\d{2}-\d{2}", path.stem)
    if len(dates) < 2:
        raise ValueError(f"文件名必须包含开始和结束日期：{path.name}")
    start, end = dates[0], dates[-1]
    start_date = datetime.strptime(start, "%Y-%m-%d")
    end_date = datetime.strptime(end, "%Y-%m-%d")
    days_covered = (end_date - start_date).days + 1
    days_in_month = monthrange(end_date.year, end_date.month)[1]
    is_partial = mode == "monthly" and (
        start_date.day != 1 or end_date.day != days_in_month
    )
    if mode == "monthly":
        label = f"{end_date.year}年{end_date.month}月"
        if is_partial:
            label += f"（截至{end_date.day}日）"
    else:
        label = f"{start[5:]} ~ {end[5:]}"
    return {
        "start": start,
        "end": end,
        "label": label,
        "daysCovered": days_covered,
        "daysInMonth": days_in_month,
        "isPartial": is_partial,
    }


def read_report(path: Path) -> OrderedDict[str, dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    missing = [column for column in REQUIRED_COLUMNS if column not in headers]
    if missing:
        raise ValueError(f"{path.name} 缺少列：{'、'.join(missing)}")
    indexes = {column: headers.index(column) for column in REQUIRED_COLUMNS}
    companies: OrderedDict[str, dict] = OrderedDict()

    for row_number, row in enumerate(rows, start=2):
        company_value = row[indexes["客户公司"]]
        customer_value = row[indexes["客户名称"]]
        if company_value is None and customer_value is None:
            continue
        if company_value is None or customer_value is None:
            raise ValueError(f"{path.name} 第 {row_number} 行缺少客户公司或客户名称")

        company = str(company_value).strip()
        customer = str(customer_value).strip()
        sales = number(row[indexes["含税金额"]])
        margin_amount = number(row[indexes["参考毛利额"]])

        company_bucket = companies.setdefault(company, {
            "sales": 0.0,
            "marginAmount": 0.0,
            "customers": OrderedDict(),
        })
        customer_bucket = company_bucket["customers"].setdefault(customer, {
            "sales": 0.0,
            "marginAmount": 0.0,
        })
        company_bucket["sales"] += sales
        company_bucket["marginAmount"] += margin_amount
        customer_bucket["sales"] += sales
        customer_bucket["marginAmount"] += margin_amount

    workbook.close()
    return companies


def rounded(value: float) -> float:
    return round(value + 1e-9, 2)


def build_companies(previous: OrderedDict[str, dict], current: OrderedDict[str, dict]) -> list[dict]:
    names = set(previous) | set(current)
    ordered_names = sorted(
        names,
        key=lambda name: (
            -current.get(name, {}).get("sales", 0),
            -previous.get(name, {}).get("sales", 0),
            name,
        ),
    )
    result = []

    for company_name in ordered_names:
        w1 = previous.get(company_name, {"sales": 0, "marginAmount": 0, "customers": {}})
        w2 = current.get(company_name, {"sales": 0, "marginAmount": 0, "customers": {}})
        customer_names = set(w1["customers"]) | set(w2["customers"])
        customers = []
        for customer_name in sorted(
            customer_names,
            key=lambda name: (
                -w2["customers"].get(name, {}).get("sales", 0),
                -w1["customers"].get(name, {}).get("sales", 0),
                name,
            ),
        ):
            c1 = w1["customers"].get(customer_name, {"sales": 0, "marginAmount": 0})
            c2 = w2["customers"].get(customer_name, {"sales": 0, "marginAmount": 0})
            customers.append({
                "name": customer_name,
                "w1Sales": rounded(c1["sales"]),
                "w2Sales": rounded(c2["sales"]),
                "w1MarginAmount": rounded(c1["marginAmount"]),
                "w2MarginAmount": rounded(c2["marginAmount"]),
            })

        result.append({
            "company": company_name,
            "w1Sales": rounded(w1["sales"]),
            "w2Sales": rounded(w2["sales"]),
            "w1MarginAmount": rounded(w1["marginAmount"]),
            "w2MarginAmount": rounded(w2["marginAmount"]),
            "customers": customers,
        })
    return result


def period_summary(period: dict[str, object], companies: OrderedDict[str, dict]) -> dict:
    sales = rounded(sum(item["sales"] for item in companies.values()))
    margin_amount = rounded(sum(item["marginAmount"] for item in companies.values()))
    return {
        **period,
        "sales": sales,
        "marginAmount": margin_amount,
        "marginRate": round(margin_amount / sales, 6) if sales else 0,
        "companyCount": sum(item["sales"] > 0 for item in companies.values()),
        "customerCount": sum(
            customer["sales"] > 0
            for item in companies.values()
            for customer in item["customers"].values()
        ),
    }


def load_history(history_path: Path) -> list[dict]:
    if not history_path.exists():
        return []
    return json.loads(history_path.read_text(encoding="utf-8")).get("periods", [])


def save_json(path: Path, payload: dict) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def main() -> None:
    mode = "weekly"
    args = sys.argv[1:]
    if len(args) >= 2 and args[0] == "--mode":
        mode = args[1]
        args = args[2:]
    if mode not in {"weekly", "monthly"} or len(args) != 2:
        raise SystemExit(
            "用法：python scripts/update-dashboard-data.py [--mode weekly|monthly] 上期报表.xlsx 本期报表.xlsx"
        )

    data_path = MONTHLY_DATA_PATH if mode == "monthly" else WEEKLY_DATA_PATH
    history_path = MONTHLY_HISTORY_PATH if mode == "monthly" else WEEKLY_HISTORY_PATH

    previous_path = Path(args[0]).resolve()
    current_path = Path(args[1]).resolve()
    for path in (previous_path, current_path):
        if not path.exists():
            raise FileNotFoundError(path)

    previous_period = period_from_filename(previous_path, mode)
    current_period = period_from_filename(current_path, mode)
    if previous_period["start"] >= current_period["start"]:
        raise ValueError("两份报表顺序不正确：请先传上期，再传本期")

    previous = read_report(previous_path)
    current = read_report(current_path)
    companies = build_companies(previous, current)

    dashboard_payload = {
        "mode": mode,
        "generatedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
        "previous": previous_period,
        "current": current_period,
        "companies": companies,
    }
    save_json(data_path, dashboard_payload)

    history_by_period = {
        f"{item['start']}|{item['end']}": item
        for item in load_history(history_path)
    }
    for period, data in ((previous_period, previous), (current_period, current)):
        history_by_period[f"{period['start']}|{period['end']}"] = period_summary(period, data)
    history = sorted(history_by_period.values(), key=lambda item: item["start"])[-12:]
    save_json(history_path, {"periods": history})

    current_summary = history_by_period[f"{current_period['start']}|{current_period['end']}"]
    print(
        f"已更新：{current_period['label']}，"
        f"{current_summary['companyCount']} 家公司，"
        f"{current_summary['customerCount']} 个客户，"
        f"销售额 {current_summary['sales']:.2f}，"
        f"毛利额 {current_summary['marginAmount']:.2f}"
    )


if __name__ == "__main__":
    main()
